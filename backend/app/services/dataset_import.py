# xlsx 헤더를 표준 컬럼으로 매핑하고 정규화 모델로 변환하는 서비스
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.domain.dataset import (
    NormalizedExpense,
    NormalizedOnlineSale,
    NormalizedSale,
)
from app.domain.enums import (
    DatasetFileType,
    DatasetFormat,
    ExpenseCategory,
    OnlineSalesReconciliationType,
)


class WorkbookProcessingError(ValueError):
    pass


@dataclass(frozen=True)
class FileSchema:
    detected_format: DatasetFormat
    aliases: dict[str, tuple[str, ...]]
    required_columns: tuple[str, ...]


@dataclass(frozen=True)
class ParsedWorkbook:
    file_type: DatasetFileType
    detected_format: DatasetFormat
    column_mapping: dict[str, str]
    missing_columns: list[str]
    mapping_confidence: float
    rows: list[dict[str, Any]]


API_COLUMN_NAMES = {
    "business_date": "businessDate",
    "transaction_time": "transactionTime",
    "receipt_number": "receiptNumber",
    "pos_number": "posNumber",
    "gross_sales": "grossSales",
    "discount_amount": "discountAmount",
    "refund_amount": "refundAmount",
    "net_sales": "netSales",
    "payment_method": "paymentMethod",
    "transaction_status": "transactionStatus",
    "transaction_date": "transactionDate",
    "counterparty": "counterparty",
    "description": "description",
    "expense_category": "expenseCategory",
    "supply_amount": "supplyAmount",
    "vat_amount": "vatAmount",
    "tax_exempt_amount": "taxExemptAmount",
    "total_amount": "totalAmount",
    "evidence_type": "evidenceType",
    "sales_channel": "salesChannel",
    "order_type": "orderType",
    "order_count": "orderCount",
    "gross_order_amount": "grossOrderAmount",
    "sales_amount": "salesAmount",
    "platform_fee_amount": "platformFeeAmount",
    "payment_fee_amount": "paymentFeeAmount",
    "merchant_delivery_fee": "merchantDeliveryFee",
    "settlement_amount": "settlementAmount",
    "settlement_date": "settlementDate",
    "settlement_status": "settlementStatus",
    "reconciliation_type": "reconciliationType",
}

FILE_SCHEMAS = {
    DatasetFileType.SALE: FileSchema(
        detected_format=DatasetFormat.EASYPOS_SALES,
        aliases={
            "business_date": ("영업일자", "거래일자", "매출일자", "business_date", "date"),
            "transaction_time": ("거래시간", "결제시간", "transaction_time"),
            "receipt_number": ("영수증번호", "거래번호", "receipt_number"),
            "pos_number": ("포스번호", "pos번호", "pos_number"),
            "gross_sales": ("총매출", "총매출액", "gross_sales"),
            "discount_amount": ("할인금액", "할인액", "discount_amount"),
            "refund_amount": ("환불금액", "반품금액", "refund_amount"),
            "net_sales": ("순매출", "순매출액", "실매출", "net_sales"),
            "payment_method": ("결제수단", "payment_method"),
            "transaction_status": ("거래상태", "transaction_status"),
        },
        required_columns=("business_date", "net_sales"),
    ),
    DatasetFileType.EXPENSE: FileSchema(
        detected_format=DatasetFormat.EASYSHOP_EXPENSE_LEDGER,
        aliases={
            "transaction_date": ("거래일자", "비용일자", "transaction_date", "date"),
            "counterparty": ("거래처", "counterparty"),
            "description": ("적요", "내용", "description"),
            "expense_category": ("비용항목", "비용분류", "expense_category"),
            "supply_amount": ("공급가액", "supply_amount"),
            "vat_amount": ("부가세", "vat", "vat_amount"),
            "tax_exempt_amount": ("면세금액", "tax_exempt_amount"),
            "total_amount": ("합계금액", "총비용", "total_amount"),
            "payment_method": ("결제수단", "payment_method"),
            "evidence_type": ("증빙유형", "evidence_type"),
        },
        required_columns=("transaction_date", "expense_category", "total_amount"),
    ),
    DatasetFileType.ONLINE_SALE: FileSchema(
        detected_format=DatasetFormat.EASYSHOP_ONLINE_SALES,
        aliases={
            "business_date": ("영업일자", "주문일자", "business_date", "date"),
            "sales_channel": ("판매채널", "플랫폼", "sales_channel"),
            "order_type": ("주문유형", "order_type"),
            "order_count": ("주문건수", "order_count"),
            "gross_order_amount": ("총주문금액", "gross_order_amount"),
            "sales_amount": ("매출금액", "순매출액", "sales_amount"),
            "discount_amount": ("할인금액", "discount_amount"),
            "refund_amount": ("환불금액", "취소환불금액", "refund_amount"),
            "platform_fee_amount": ("플랫폼수수료", "platform_fee_amount"),
            "payment_fee_amount": ("결제수수료", "payment_fee_amount"),
            "merchant_delivery_fee": ("점주부담배달비", "merchant_delivery_fee"),
            "settlement_amount": ("정산금액", "settlement_amount"),
            "settlement_date": ("정산일자", "settlement_date"),
            "settlement_status": ("정산상태", "settlement_status"),
            "reconciliation_type": ("대사유형", "reconciliation_type"),
        },
        required_columns=("business_date", "sales_amount"),
    ),
}


def analyze_workbook(contents: bytes, file_type: DatasetFileType) -> ParsedWorkbook:
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise WorkbookProcessingError("xlsx 파일을 읽을 수 없습니다.") from error

    try:
        worksheet = workbook.active
        worksheet_rows = list(worksheet.iter_rows(values_only=True))
        if not worksheet_rows:
            raise WorkbookProcessingError("xlsx 파일에 데이터가 없습니다.")

        schema = FILE_SCHEMAS[file_type]
        header_index, indexed_mapping = _find_header_row(worksheet_rows, schema)

        mapped_columns = set(indexed_mapping.values())
        missing_columns = [
            API_COLUMN_NAMES[column]
            for column in schema.required_columns
            if column not in mapped_columns
        ]
        column_mapping = {
            str(worksheet_rows[header_index][index]).strip(): API_COLUMN_NAMES[column]
            for index, column in indexed_mapping.items()
        }
        data_rows = _extract_rows(worksheet_rows[header_index + 1 :], indexed_mapping)

        return ParsedWorkbook(
            file_type=file_type,
            detected_format=(schema.detected_format if indexed_mapping else DatasetFormat.UNKNOWN),
            column_mapping=column_mapping,
            missing_columns=missing_columns,
            mapping_confidence=round(len(mapped_columns) / len(schema.aliases), 4),
            rows=data_rows,
        )
    finally:
        workbook.close()


def normalize_rows(
    parsed: ParsedWorkbook,
    dataset_id: int,
    source_file_id: int,
) -> list[NormalizedSale | NormalizedExpense | NormalizedOnlineSale]:
    if parsed.missing_columns:
        return []

    try:
        if parsed.file_type == DatasetFileType.SALE:
            return [
                NormalizedSale(
                    dataset_id=dataset_id,
                    source_file_id=source_file_id,
                    business_date=_to_date(row.get("business_date")),
                    transaction_time=_to_datetime(row.get("transaction_time")),
                    receipt_number=_to_optional_text(row.get("receipt_number")),
                    pos_number=_to_optional_text(row.get("pos_number")),
                    gross_sales=_to_int(row.get("gross_sales")),
                    discount_amount=_to_int(row.get("discount_amount")),
                    refund_amount=_to_int(row.get("refund_amount")),
                    net_sales=_to_int(row.get("net_sales")),
                    payment_method=_to_optional_text(row.get("payment_method")),
                    transaction_status=_to_optional_text(row.get("transaction_status")),
                )
                for row in parsed.rows
            ]
        if parsed.file_type == DatasetFileType.EXPENSE:
            return [
                NormalizedExpense(
                    dataset_id=dataset_id,
                    source_file_id=source_file_id,
                    transaction_date=_to_date(row.get("transaction_date")),
                    counterparty=_to_optional_text(row.get("counterparty")),
                    description=_to_optional_text(row.get("description")),
                    expense_category=_to_expense_category(row.get("expense_category")),
                    supply_amount=_to_int(row.get("supply_amount")),
                    vat_amount=_to_int(row.get("vat_amount")),
                    tax_exempt_amount=_to_int(row.get("tax_exempt_amount")),
                    total_amount=_to_int(row.get("total_amount")),
                    payment_method=_to_optional_text(row.get("payment_method")),
                    evidence_type=_to_optional_text(row.get("evidence_type")),
                )
                for row in parsed.rows
            ]
        return [
            NormalizedOnlineSale(
                dataset_id=dataset_id,
                source_file_id=source_file_id,
                business_date=_to_date(row.get("business_date")),
                sales_channel=_to_optional_text(row.get("sales_channel")),
                order_type=_to_optional_text(row.get("order_type")),
                order_count=_to_int(row.get("order_count")),
                gross_order_amount=_to_int(row.get("gross_order_amount")),
                sales_amount=_to_int(row.get("sales_amount")),
                discount_amount=_to_int(row.get("discount_amount")),
                refund_amount=_to_int(row.get("refund_amount")),
                platform_fee_amount=_to_int(row.get("platform_fee_amount")),
                payment_fee_amount=_to_int(row.get("payment_fee_amount")),
                merchant_delivery_fee=_to_int(row.get("merchant_delivery_fee")),
                settlement_amount=_to_optional_int(row.get("settlement_amount")),
                settlement_date=_to_optional_date(row.get("settlement_date")),
                settlement_status=_to_optional_text(row.get("settlement_status")),
                reconciliation_type=_to_reconciliation_type(row.get("reconciliation_type")),
            )
            for row in parsed.rows
        ]
    except (TypeError, ValueError) as error:
        raise WorkbookProcessingError("xlsx 행을 정규화할 수 없습니다.") from error


def _find_header_row(
    rows: list[tuple[Any, ...]],
    schema: FileSchema,
) -> tuple[int, dict[int, str]]:
    aliases = {
        _normalize_header(alias): column
        for column, column_aliases in schema.aliases.items()
        for alias in column_aliases
    }
    best_index = 0
    best_mapping: dict[int, str] = {}
    for index, row in enumerate(rows[:10]):
        mapping: dict[int, str] = {}
        used_columns: set[str] = set()
        for column_index, header in enumerate(row):
            canonical = aliases.get(_normalize_header(header))
            if canonical and canonical not in used_columns:
                mapping[column_index] = canonical
                used_columns.add(canonical)
        if len(mapping) > len(best_mapping):
            best_index = index
            best_mapping = mapping
    return best_index, best_mapping


def _extract_rows(
    rows: list[tuple[Any, ...]],
    indexed_mapping: dict[int, str],
) -> list[dict[str, Any]]:
    extracted = []
    for row in rows:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        extracted.append(
            {
                canonical: row[index] if index < len(row) else None
                for index, canonical in indexed_mapping.items()
            }
        )
    return extracted


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("_", "")


def _to_date(value: object) -> date:
    parsed = _to_optional_date(value)
    if parsed is None:
        raise ValueError("날짜 값이 필요합니다.")
    return parsed


def _to_optional_date(value: object) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for date_format in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError("지원하지 않는 날짜 형식입니다.")


def _to_datetime(value: object) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return datetime.fromisoformat(str(value).strip())


def _to_int(value: object) -> int:
    if value is None or str(value).strip() == "":
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(float(str(value).strip().replace(",", "").replace("원", "")))


def _to_optional_int(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    return _to_int(value)


def _to_optional_text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _to_expense_category(value: object) -> ExpenseCategory:
    normalized = _normalize_header(value)
    categories = {
        "재료비": ExpenseCategory.MATERIAL,
        "material": ExpenseCategory.MATERIAL,
        "인건비": ExpenseCategory.LABOR,
        "labor": ExpenseCategory.LABOR,
        "임차료": ExpenseCategory.RENT,
        "임대료": ExpenseCategory.RENT,
        "rent": ExpenseCategory.RENT,
        "공과금": ExpenseCategory.UTILITY,
        "utility": ExpenseCategory.UTILITY,
        "광고비": ExpenseCategory.MARKETING,
        "마케팅": ExpenseCategory.MARKETING,
        "marketing": ExpenseCategory.MARKETING,
        "수수료": ExpenseCategory.COMMISSION,
        "commission": ExpenseCategory.COMMISSION,
        "플랫폼수수료": ExpenseCategory.PLATFORM_FEE,
        "platformfee": ExpenseCategory.PLATFORM_FEE,
        "소모품": ExpenseCategory.SUPPLIES,
        "supplies": ExpenseCategory.SUPPLIES,
        "유지보수": ExpenseCategory.MAINTENANCE,
        "maintenance": ExpenseCategory.MAINTENANCE,
    }
    return categories.get(normalized, ExpenseCategory.OTHER)


def _to_reconciliation_type(value: object) -> OnlineSalesReconciliationType:
    normalized = _normalize_header(value)
    if normalized in {"separatefrompostotal", "별도매출", "포스매출별도"}:
        return OnlineSalesReconciliationType.SEPARATE_FROM_POS_TOTAL
    return OnlineSalesReconciliationType.INCLUDED_IN_POS_TOTAL
